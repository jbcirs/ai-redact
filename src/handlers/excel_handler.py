#!/usr/bin/env python3
"""
excel_handler.py — Kind B native handler for Excel workbooks (.xlsx).

Redacts IN the xlsx format (docs/plans/format-support-plan.md §3.2).
Output is VALUES-ONLY by design: the workbook is loaded with
data_only=True so every formula is replaced by its cached computed value.
Keeping formulas would be strictly worse — `=A1&B1` could reconstruct a
redacted cell. This is stated loudly in the report notes.

Scanned (and redacted) surfaces:
  - every cell value on EVERY sheet, including hidden sheets
  - each data cell additionally scanned with its column header as context
    ("<header>: <value>") so contextual patterns can fire; replacements
    are only ever applied within the cell's own value
  - cell comments
  - sheet titles (renamed when matched)
  - defined names (dropped entirely when matched — counted)
  - workbook core properties (cleared unconditionally — noted)

Drawing parts (shapes/text boxes/charts) are not carried over by this
values-focused path: their presence is detected in the raw zip and
counted as dropped — dropped content cannot leak, but it is never silent.

Local only. No network. No subprocesses.
"""

from __future__ import annotations

import sys
import zipfile
from pathlib import Path

try:  # imported as part of the handlers package (normal router path)
    from handlers.common import UnsupportedFormatError, redaction_text
except ImportError:  # executed directly: python src/handlers/excel_handler.py
    from common import UnsupportedFormatError, redaction_text

SUPPORTED_EXTENSIONS = {".xlsx"}

# Core-property fields that can carry user text/PII.
_CORE_PROPERTY_FIELDS = (
    "creator",
    "title",
    "subject",
    "description",
    "keywords",
    "lastModifiedBy",
    "category",
    "contentStatus",
    "identifier",
    "language",
    "version",
)


def _load_workbook(path: Path, data_only: bool = True):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UnsupportedFormatError(
            "openpyxl is not installed — run "
            "'pip install -r requirements.txt' to process .xlsx files."
        ) from exc
    try:
        return load_workbook(str(path), data_only=data_only)
    except Exception as exc:
        raise UnsupportedFormatError(
            f"could not open {Path(path).name} as an .xlsx workbook "
            f"({exc}). Legacy .xls files must be saved as .xlsx first; "
            "password-protected workbooks are not supported."
        ) from exc


def _redact_spans(text: str, found: list[tuple[str, str]]) -> str:
    """Replace every matched span inside `text`. Longer matches first so a
    short match never splits a longer one. A cell whose entire value
    matched naturally becomes just the redaction string."""
    for _category, matched in sorted(
        set(found), key=lambda cm: len(cm[1]), reverse=True
    ):
        text = text.replace(matched, redaction_text(matched))
    return text


def _scan_cell(text: str, header: str | None, matcher) -> list[tuple[str, str]]:
    """Scan a cell value bare AND with its column header as context, so
    contextual patterns ("Account #: 12345") can fire on labeled columns.
    Only matches that actually occur inside the cell's own value are
    returned — the header itself is never redacted through a data cell.
    Duplicate (category, match) pairs from the two scans count once."""
    found = {(cat, m) for cat, m in matcher(text) if m in text}
    if header:
        for cat, m in matcher(f"{header}: {text}"):
            if m in text:
                found.add((cat, m))
    return sorted(found)


def _sheet_headers(ws) -> dict[int, str]:
    """Row 1 of each sheet is treated as column headers (context only)."""
    headers: dict[int, str] = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
        if cell.value is not None:
            text = str(cell.value).strip()
            if text:
                headers[cell.column] = text
    return headers


def _count_uncached_formulas(input_path: Path, wb_values) -> int:
    """Double-load audit: formula cells whose cached value is missing come
    out EMPTY in the values-only output. Count them so the loss is never
    silent."""
    try:
        wb_formulas = _load_workbook(input_path, data_only=False)
    except UnsupportedFormatError:
        raise
    except Exception:
        return 0
    uncached = 0
    for ws_f in wb_formulas.worksheets:
        try:
            ws_v = wb_values[ws_f.title]
        except KeyError:
            continue
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    if ws_v.cell(row=cell.row, column=cell.column).value is None:
                        uncached += 1
    wb_formulas.close()
    return uncached


def _count_drawing_parts(input_path: Path) -> int:
    """Detect drawing parts (shapes/text boxes/charts) in the raw zip.
    openpyxl does not carry them into the rewritten workbook — dropped,
    not leaked, but always counted."""
    try:
        with zipfile.ZipFile(input_path) as zf:
            return sum(
                1
                for name in zf.namelist()
                if name.startswith("xl/drawings/") and name.endswith(".xml")
            )
    except Exception:
        return 0


def _iter_defined_names(wb):
    """Yield (scope_label, container, name, defined_name) for workbook- and
    worksheet-scoped defined names (openpyxl >= 3.1 dict-style API)."""
    for name, dn in list(wb.defined_names.items()):
        yield "workbook", wb.defined_names, name, dn
    for ws in wb.worksheets:
        ws_names = getattr(ws, "defined_names", None)
        if ws_names:
            for name, dn in list(ws_names.items()):
                yield f"sheet {ws.title}", ws_names, name, dn


# ---------------------------------------------------------------------------
# Kind B contract: redact_file / verify_file
# ---------------------------------------------------------------------------


def redact_file(
    input_path,
    output_path,
    matcher,
    dry_run: bool,
    options: dict,
) -> dict:
    """Scan every surface of the workbook; unless dry_run, write the
    redacted values-only copy to output_path. Never touches the input."""
    input_path = Path(input_path)
    wb = _load_workbook(input_path, data_only=True)

    counts: dict[str, int] = {}
    matches: list[tuple[str, str, str, bool]] = []
    notes: list[str] = [
        "formulas replaced with computed values (values-only output)"
    ]
    unit_count = 0

    def record(unit: str, found: list[tuple[str, str]]) -> None:
        for category, matched in found:
            counts[category] = counts.get(category, 0) + 1
            matches.append((unit, category, matched, True))

    hidden_sheets = [
        ws.title for ws in wb.worksheets if ws.sheet_state != "visible"
    ]
    if hidden_sheets:
        notes.append(
            "hidden sheet(s) scanned: " + ", ".join(hidden_sheets)
        )

    # --- cells + comments, every sheet including hidden ---------------------
    for ws in wb.worksheets:
        headers = _sheet_headers(ws)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    unit_count += 1
                    text = str(cell.value)
                    header = headers.get(cell.column) if cell.row > 1 else None
                    found = _scan_cell(text, header, matcher)
                    if found:
                        record(f"{ws.title}!{cell.coordinate}", found)
                        if not dry_run:
                            cell.value = _redact_spans(text, found)
                if cell.comment is not None:
                    comment_text = cell.comment.text or ""
                    found = _scan_cell(comment_text, None, matcher)
                    if found:
                        record(
                            f"{ws.title}!{cell.coordinate} (comment)", found
                        )
                        if not dry_run:
                            from openpyxl.comments import Comment

                            cell.comment = Comment(
                                _redact_spans(comment_text, found),
                                cell.comment.author or "",
                            )

    # --- sheet titles --------------------------------------------------------
    for ws in list(wb.worksheets):
        found = _scan_cell(ws.title, None, matcher)
        if found:
            record(f"[sheet title] {ws.title}", found)
            if not dry_run:
                new_title = _redact_spans(ws.title, found)[:31]
                old_title = ws.title
                try:
                    ws.title = new_title
                    notes.append(
                        f'sheet renamed: "{old_title}" -> "{ws.title}"'
                    )
                except Exception:
                    ws.title = f"Sheet_{wb.worksheets.index(ws) + 1}"
                    notes.append(
                        f'sheet renamed: "{old_title}" -> "{ws.title}"'
                    )

    # --- defined names (dropped entirely when matched) -----------------------
    dropped_names = 0
    for scope, container, name, dn in _iter_defined_names(wb):
        value = getattr(dn, "attr_text", "") or ""
        found = _scan_cell(f"{name} {value}", None, matcher)
        if found:
            record(f"[defined name/{scope}] {name}", found)
            dropped_names += 1
            if not dry_run:
                try:
                    del container[name]
                except Exception:
                    pass
    if dropped_names:
        notes.append(
            f"{dropped_names} defined name(s) containing matches removed"
        )

    # --- core properties: scan for the report, then clear unconditionally ---
    props = wb.properties
    for field in _CORE_PROPERTY_FIELDS:
        value = getattr(props, field, None)
        if isinstance(value, str) and value:
            found = _scan_cell(value, None, matcher)
            if found:
                record(f"[core property] {field}", found)
        if not dry_run and isinstance(value, str):
            setattr(props, field, None)
    notes.append(
        "workbook core properties (creator/title/etc.) cleared"
        if not dry_run
        else "workbook core properties (creator/title/etc.) will be "
        "cleared on a real run"
    )

    # --- loss accounting (never silent) --------------------------------------
    uncached = _count_uncached_formulas(input_path, wb)
    if uncached:
        notes.append(
            f"{uncached} formula cells had no cached value "
            "(empty in output)"
        )
    dropped_elements = _count_drawing_parts(input_path)
    if dropped_elements:
        notes.append(
            f"{dropped_elements} drawing part(s) (shapes/text boxes/"
            "charts) not carried over — content dropped, not leaked"
        )

    if not dry_run:
        wb.save(str(output_path))
    wb.close()

    return {
        "counts": counts,
        "matches": matches,
        "unit_label": "cell",
        "unit_count": unit_count,
        "notes": notes,
        "dropped_elements": dropped_elements,
    }


def verify_file(output_path, matcher, options: dict) -> dict:
    """Re-open the WRITTEN workbook and re-scan every surface with the
    same construction used for redaction (including header-as-context).
    Returns {category: remaining_count} — {} means clean."""
    wb = _load_workbook(Path(output_path), data_only=True)
    remaining: dict[str, int] = {}

    def record(found: list[tuple[str, str]]) -> None:
        for category, _matched in found:
            remaining[category] = remaining.get(category, 0) + 1

    for ws in wb.worksheets:
        headers = _sheet_headers(ws)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    header = headers.get(cell.column) if cell.row > 1 else None
                    record(_scan_cell(str(cell.value), header, matcher))
                if cell.comment is not None:
                    record(_scan_cell(cell.comment.text or "", None, matcher))
        record(_scan_cell(ws.title, None, matcher))

    for _scope, _container, name, dn in _iter_defined_names(wb):
        value = getattr(dn, "attr_text", "") or ""
        record(_scan_cell(f"{name} {value}", None, matcher))

    props = wb.properties
    for field in _CORE_PROPERTY_FIELDS:
        value = getattr(props, field, None)
        if isinstance(value, str) and value:
            record(_scan_cell(value, None, matcher))

    wb.close()
    return remaining


# ---------------------------------------------------------------------------
# Standalone smoke test
# ---------------------------------------------------------------------------


def _smoke_test() -> int:
    import re
    import tempfile

    sys.path.insert(
        0, str(Path(__file__).resolve().parents[2] / "tests")
    )
    from make_office_fixtures import PLANTED_EMAIL, SURVIVOR, make_xlsx

    email_re = re.compile(
        r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    )

    def matcher(text: str) -> list[tuple[str, str]]:
        return [("email", m) for m in email_re.findall(text)]

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        xlsx_path = make_xlsx(tmp_dir)
        out_path = tmp_dir / "fixture_redacted.xlsx"

        # Dry run: scan only, no output written.
        dry = redact_file(xlsx_path, out_path, matcher, True, {})
        assert not out_path.exists(), "dry_run must not write output"
        assert dry["counts"].get("email", 0) >= 4, (
            f"dry_run undercounted: {dry['counts']}"
        )

        result = redact_file(xlsx_path, out_path, matcher, False, {})
        assert out_path.exists(), "redacted output missing"
        assert result["unit_label"] == "cell"
        assert result["counts"].get("email", 0) >= 4, (
            f"undercounted: {result['counts']}"
        )
        units = [m[0] for m in result["matches"]]
        assert any(u.startswith("Archive!") for u in units), (
            "hidden sheet was not scanned"
        )
        print(f"redact counts: {result['counts']}")
        print(f"matches: {units}")
        for note in result["notes"]:
            print(f"note: {note}")

        remaining = verify_file(out_path, matcher, {})
        assert remaining == {}, f"verify found remaining PII: {remaining}"

        # The financial string must survive; the email must be gone.
        from openpyxl import load_workbook

        wb = load_workbook(str(out_path), data_only=True)
        all_values = [
            str(cell.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        wb.close()
        assert any(SURVIVOR in v for v in all_values), (
            f"{SURVIVOR} did not survive redaction"
        )
        assert not any(PLANTED_EMAIL in v for v in all_values), (
            "planted email survived in a cell"
        )

    print("excel_handler smoke test: PASS")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(_smoke_test())
    except AssertionError as exc:
        print(f"excel_handler smoke test: FAIL — {exc}", file=sys.stderr)
        sys.exit(1)
