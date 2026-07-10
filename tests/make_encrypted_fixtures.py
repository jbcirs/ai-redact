#!/usr/bin/env python3
"""
make_encrypted_fixtures.py — generate password-protected fixtures for
handler tests (docs/plans/expansion-plan.md §3.G).

Writes into the directory given as argv[1]:
  protected.pdf   — AES-256 encrypted, user password TEST_PASSWORD
  protected.xlsx  — msoffcrypto-encrypted, password TEST_PASSWORD

Both plant PLANTED_EMAIL + the must-survive SURVIVOR string. Password is
exported as TEST_PASSWORD so scripts/test.sh can use it without embedding
a literal secret-looking string in two places.

Pure Python (PyMuPDF + msoffcrypto-tool, both already required deps), no
network.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

PLANTED_EMAIL = "planted.email@example.com"
SURVIVOR = "$12,345.67"
TEST_PASSWORD = "TestPass123!"


def make_encrypted_pdf(out_dir: Path) -> Path:
    import fitz

    path = Path(out_dir) / "protected.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 80), f"Contact: {PLANTED_EMAIL}\nBalance: {SURVIVOR}")
    doc.save(str(path), encryption=fitz.PDF_ENCRYPT_AES_256,
             user_pw=TEST_PASSWORD, owner_pw=TEST_PASSWORD + "_owner")
    doc.close()
    print(f"wrote {path}")
    return path


def make_encrypted_xlsx(out_dir: Path) -> Path:
    import msoffcrypto
    from openpyxl import Workbook

    path = Path(out_dir) / "protected.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Contact"
    ws["A2"] = PLANTED_EMAIL
    ws["A3"] = "Balance"
    ws["A4"] = SURVIVOR
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    enc = io.BytesIO()
    msoffcrypto.OfficeFile(buf).encrypt(TEST_PASSWORD, enc)
    path.write_bytes(enc.getvalue())
    print(f"wrote {path}")
    return path


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: make_encrypted_fixtures.py <output-dir>", file=sys.stderr)
        return 2
    out_dir = Path(argv[1])
    out_dir.mkdir(parents=True, exist_ok=True)
    make_encrypted_pdf(out_dir)
    make_encrypted_xlsx(out_dir)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
